import requests, openpyxl, time
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment

wb = openpyxl.load_workbook('list.xlsx', data_only = True)
ws = wb['Sheet1']

for i in range(2, 201):
    # 공백을 만나면, 프로그램 종료
    if ws.cell(i, 1).value is None: break
    nickname = ws.cell(i,1).value
    webpage = requests.get('https://maple.gg/u/' + nickname)
    soup = BeautifulSoup(webpage.content, 'html.parser')
    
    simple_info = soup.find_all(attrs={'class':'user-summary-item'})

    # 닉변해서 검색 기록이 안 나오는 캐릭터에 대한 예외 처리
    if not simple_info:
        ws.cell(i, 2).value = 'N/A'
        ws.cell(i, 3).value = 'N/A'
        ws.cell(i, 4).value = 'N/A'
        ws.cell(i, 5).value = 'N/A'
        ws.cell(i, 6).value = 'N/A'
        continue
    
    ws.cell(row=i, column=2).value = simple_info[0].get_text()
    ws.cell(i, 3).value = simple_info[1].get_text()

    highest_floor = soup.find_all(attrs={'class':'character-card-additional-item'})

    # 무릉 기록이 없는 캐릭터에 대한 if문 (예외 처리)
    if highest_floor[0].get_text() == '''
무릉도장
기록없음
''':
        ws.cell(i, 4).value = 'N/A'
        ws.cell(i, 5).value = 'N/A'
        ws.cell(i, 6).value = 'N/A'
        continue
    
    ws.cell(i, 4).value = highest_floor[0].get_text()[9:12]
    ws.cell(i, 5).value = highest_floor[0].get_text()[12:-1]

    floor_renewal_date = soup.find_all(attrs={'class':'user-summary-date'})
    ws.cell(i, 6).value = floor_renewal_date[0].get_text()[6:-1]

for i in range(1, 201):
    for j in range(1, 7):
        ws.cell(i, j).alignment = Alignment(horizontal = 'center')


print('Work is Done, Open your Excel File')
wb.save('list.xlsx')
